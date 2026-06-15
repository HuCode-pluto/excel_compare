# -*- coding: utf-8 -*-
import pandas as pd
import sys
import os
import traceback
from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# 自定义输出类：同时输出到终端和日志文件
class TeeOutput:
    def __init__(self, terminal, log):
        self.terminal = terminal
        self.log = log
    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()
    def flush(self):
        self.terminal.flush()
        self.log.flush()


# ========== 通用转换&清洗函数 ==========
# def full_clean(s):
#     """
#     深度清洗：去除所有空格、制表符、全角空格、空白字符，统一为纯字符串
#     解决肉眼一致、隐形字符导致的误判
#     """
#     if pd.isna(s):
#         return ""
#     s = str(s)
#     # 常规半角空格、制表符、换行
#     s = s.replace(" ", "").replace("\t", "").replace("\n", "").replace("\r", "")
#     # 全角空格
#     s = s.replace("　", "")
#     return s

def full_clean(s):
    """
    深度清洗 + 数值标准化：
    1. 清除各类空格/空白符
    2. 数字(含科学计数、1.0、超大数)统一转为纯整数字符串
    3. 非文本保留原始清洗结果
    """
    if pd.isna(s):
        return ""
    # 第一步：先清除所有空白字符（原有逻辑保留）
    s = str(s)
    s = s.replace(" ", "").replace("\t", "").replace("\n", "").replace("\r", "")
    s = s.replace("　", "")

    # 第二步：新增【数值统一格式化】（解决 1/1.0、科学计数 问题）
    try:
        # 尝试转为高精度数字，兼容科学计数、小数
        num = Decimal(s)
        # 如果是整数，强制转为整数字符串（去掉 .0、科学计数）
        if num == num.to_integral_value():
            return str(num.to_integral_value())
        # 非整数浮点数，保留原值（按需可自行扩展）
        return str(num)
    except (InvalidOperation, ValueError):
        # 不是数字（普通中文、字母、混合文本），直接返回清洗后字符串
        return s


def trim_str_length(s, max_length):
    if pd.isna(s):
        return ""
    s = full_clean(s)
    result = str(s).strip()
    if max_length is not None:
        result = result[:max_length]
    return result


def left_trim_by_ref_len(target_col, ref_col):
    """
    通用跨列处理函数：目标列内容从左侧去除【参考列的字符长度】
    参数：
        target_col: 要处理的目标列名（被截断的列）
        ref_col: 参考列名（用它的长度来截断）
    返回：可直接用于行级apply的处理函数
    """

    # 新增：给处理函数标记目标列属性
    def _process_row(row):
        # 空值安全处理
        target_val = row[target_col]
        ref_val = row[ref_col]

        target_str = str(target_val).strip() if pd.notna(target_val) else ""
        ref_str = str(ref_val).strip() if pd.notna(ref_val) else ""

        ref_len = len(ref_str)
        # 目标列长度 <= 参考列长度，返回空
        if len(target_str) <= ref_len:
            return ""
        # 从左侧截断，返回剩余内容
        return target_str[ref_len:]

    # 标记目标列，方便后续提取
    _process_row.target_col = target_col
    # 标记参考列，方便后续提取
    _process_row.ref_col = ref_col
    return _process_row


def right_keep_by_ref_len(target_col, ref_col):
    """
    通用跨列处理函数：目标列内容从右侧开始，保留【参考列字符长度】的字符
    参数：
        target_col: 要处理的目标列名（被截取的列）
        ref_col: 参考列名（用它的长度决定保留多少位）
    返回：可直接用于行级apply的处理函数
    """
    def _process_row(row):
        # 空值安全处理
        target_val = row[target_col]
        ref_val = row[ref_col]

        target_str = str(target_val).strip() if pd.notna(target_val) else ""
        ref_str = str(ref_val).strip() if pd.notna(ref_val) else ""

        ref_len = len(ref_str)
        # 参考列为空 / 长度为0，返回空
        if ref_len <= 0:
            return ""
        # 目标列长度 <= 参考列长度，原样返回全部内容
        if len(target_str) <= ref_len:
            return target_str
        # 从右侧截取，保留指定长度
        return target_str[-ref_len:]

    # 标记目标列，兼容现有跨列处理逻辑
    _process_row.target_col = target_col
    # 标记参考列，方便后续提取
    _process_row.ref_col = ref_col
    return _process_row



def convert_station_type(s):
    """中文/数字 统一映射"""
    if pd.isna(s):
        return ""
    s = full_clean(s)

    mapping = {
        "杆变": "0",
        "箱式变": "1",
        "分布式电源": "2",
        "非接地刀闸": "2",
        "配网小车刀闸": "100",
        "非常重要": "1",
        "比较重要": "2",
        "一般": "3",
        "横向": "0",
        "竖向": "1",
        "支撑杆": "0",
        "断联杆": "1",
        "耐张杆": "2",
        "跨越杆": "3",
        "辅杆": "4",
        "转角杆": "5",
        "黑图": "0",
        "红图": "1",
        "黄图": "2",
        "接地刀闸": "3",
        "光纤纵差联络开关": "15",
        "出线开关": "16",
        "负荷开关": "4",
        "电缆头": "7",
        "线路开关": "8",
        "母联开关": "9",
        "常规联络开关": "10",
        "分布式电源控制开关": "11",
        "准同期开关": "12",
        "专用联络开关": "13",
        "电压型开关": "14",
        "小车开关": "99",
        "分段开关": "0",
        "联络开关": "1",
        "电磁型傻瓜开关": "0",
        "电压型智能开关": "1",
        "不可拆搭": "0",
        "可拆搭": "1",
        "架空线": "0",
        "电缆": "1",
        "专供": "2",
        "大用户": "3",
        "未变化": "0",
        "更新": "1",
        "删除": "2",
        "新增": "3",
        "投运": "0",
        "未投运": "1",
        "待退役": "2",
        "配网开关": "13502",
        "配网变压器": "13505",
        "配网母线": "13506",
        "配网刀闸": "13513",
        "配网接地刀闸": "13514",
        "配网故障指示器": "13523",
        "配网开关站": "13501",
        "配网终端": "13510",
        "正交变换法": "0",
        "正则化法": "1",
        "混合法": "2",
        "风险主题": "1",
        "运行方式": "2",
        "备用主题3": "3",
        "备用主题4": "4",
        "开关站": "0",
        "环网柜": "1",
        "街坊站": "2",
        "光伏站": "3",
        "箱式变电站": "4",
        "配电站": "5",
        "用户站": "6",
        "电缆分支箱": "7",
        "低压配变箱": "8",
        "低压电缆分支箱": "9",
        "低压电缆终端箱": "10",
        "未定义类别": "0",
        "DTU": "1",
        "FTU": "2",
        "TTU": "3",
        "配网子站": "4",
        "故障指示器": "5",
        "DTU－开关站": "6",
        "其它": "0",
        "事故总": "1",
        "预告信号": "2",
        "动作信号": "3",
        "故障信号": "4",
        "通讯信号": "5",
        "控制信号": "6",
        "设备信号": "7",
        "重合闸信号": "9",
        "备自投信号": "10",
        "机组人工停机信号": "11",
        "厂站VQC信号": "12",
        "纵差保护": "13",
        "短路故障": "19",
        "接地故障": "20",
        "电池活化": "21",
        "纵差后备保护": "22",
        "Uab": "1",
        "Ubc": "2",
        "Uca": "3",
        "Uab1": "4",
        "Ubc1": "5",
        "Uca1": "6",
        "Uab2": "7",
        "Ubc2": "8",
        "Uca2": "9",
        "Uab1/Ua1电压": "10",
        "Ucb1/Ub1电压": "11",
        "Uab2/Ua2电压": "12",
        "Ucb2/Ub2电压": "13",
        "零序电压U0": "14",
        "蓄电池电压Ue": "15",
        "零序电流I0": "16",
        "站用电Ua": "17",
        "站用电Ub": "18",
        "站用电Uc": "19",
        "交流屏Ua": "20",
        "交流屏Ub": "21",
        "交流屏Uc": "22",
        "交流屏1Ub": "23",
        "交流屏2Ub": "24",
        "直流屏U": "25",
        "直流屏控母电压": "26",
        "直流电压": "27",
        "充电电压": "28",
        "终端电压": "29",
        "装置温度": "30",
        "开口电压": "31",
        "频率": "32",
        "A相电压幅值": "33",
        "B相电压幅值": "34",
        "C相电压幅值": "35",
        "气象站总辐射": "36",
        "气象站直辐射": "37",
        "气象站散辐射": "38",
        "气象站风速": "39",
        "气象站风向": "40",
        "气象站温度": "41",
        "气象站湿度": "42",
        "气象站气压": "43",
        "PT断线": "44",
        "控回断线": "45",
        "零流方式": "46",
        "相间过流一段控制字": "47",
        "相间过流一段电流": "48",
        "相间过流一段延时": "49",
        "相间过流二段控制字": "50",
        "相间过流二段电流": "51",
        "相间过流二段延时": "52",
        "相间过流三段控制字": "53",
        "相间过流三段电流": "54",
        "相间过流三段延时": "55",
        "过流反时限控制字": "56",
        "过流反时限启动值": "57",
        "过流反时限时间": "58",
        "过负荷保护控制字": "59",
        "过负荷保护电流": "60",
        "过负荷保护延时": "61",
        "一次重合闸控制字": "62",
        "一次重合闸延时": "63",
        "加速段过流控制字": "64",
        "加速段过流电流": "65",
        "加速段过流延时": "66",
        "零序过流一段控制字": "67",
        "零序过流一段电流": "68",
        "零序过流一段延时": "69",
        "零序过流二段控制字": "70",
        "零序过流二段电流": "71",
        "零序过流二段延时": "72",
        "零流反时限控制字": "73",
        "零流反时限启动值": "74",
        "零流反时限时间": "75",
        "涌流识别控制字": "76",
        "二次谐波闭锁系数": "77",
        "PT变比": "78",
        "CT变比": "79",
        "零序PT变比": "80",
        "零序CT变比": "81",
        "零流后加速投入控制字": "82",
        "零流后加速定值": "83",
        "零流后加速时间": "84",
        "FA模式控制字": "85",
        "二次重合闸投入控制字": "86",
        "二次重合闸闭锁控制字": "87",
        "二次重合闸延时": "88",
        "保护总投入控制字": "89",
        "分段X时间": "90",
        "分段X长延时": "91",
        "分段Y时间": "92",
        "失压分闸延时": "93",
        "联络XL时间": "94",
        "联络YL时间": "95",
        "小电流接地控制字": "96",
        "AGC控制对象有功功率": "97",
        "AGC控制对象有功可调上限": "98",
        "AGC控制对象有功可调下限": "99",
        "AGC调节指令返回值": "100",
        "AGC控制对象可增无功": "101",
        "AGC控制对象可减无功": "102",
        "AGC控制对象电压目标值": "103",
        "AGC控制对象无功目标值": "104",
        "一次调频动作实时频率": "105",
        "远方全厂AGC目标值": "106",
        "本地全厂AGC目标值": "107",
        "AVC控制对象厂站可调电压上限": "108",
        "AVC控制对象厂站可调电压下限": "109",
        "在线": "0",
        "离线": "1",
        "仿真": "2",
        "交互方式": "0",
        "自动方式": "1",
        "分闸加事故总": "0",
        "分合分": "1",
        "非正常分闸": "2",
        "分闸加保护": "3",
        "IEC(JM)104-N": "177",
        "华北103规约": "190",
        "IEC103规约扩展": "191",
        "南网103规约": "192",
        "扩展南网103规约": "193",
        "河南103规约": "194",
        "华中103规约": "195",
        "江苏103规约": "196",
        "安徽103规约": "197",
        "宁夏103规约": "198",
        "宁夏扩展103规约": "199",
        "TrueTime规约New": "200",
        "TrueTime规约Old": "201",
        "天文钟规约(NARI)": "202",
        "VCOM天文钟": "205",
        "北斗天文钟": "206",
        "YDJ天文钟": "207",
        "NARI_TSS天文钟": "208",
        "IEC-104-REG": "209",
        "CDT": "1",
        "BJ9702": "2",
        "DISA": "3",
        "NT8000": "8",
        "PH(JM)101": "20",
        "IEC-101": "21",
        "IEC-102": "22",
        "IEC-103": "23",
        "IEC-104": "24",
        "PH101": "27",
        "IEC-104程序化控制": "28",
        "IEC(JM)-104": "29",
        "U4F": "30",
        "SC1801": "31",
        "S5": "32",
        "DNP": "33",
        "CDC8890": "34",
        "IEC-104-GFJM": "37",
        "IEEE118动态管理规约": "297",
        "IEEE118动态数据规约": "298",
        "IEEE118动态文件规约": "299",
        "IEEE1344动态管理规约": "300",
        "IEEE1344动态数据规约": "301",
        "IEEE1344动态文件规约": "302",
        "天文钟规约(PMU)": "304",
        "DL476": "50",
        "DL476FILE": "52",
        "DL476alarm_send": "53",
        "DL476alarm_recv": "54",
        "TASE2": "60",
        "IEC101-POLLING": "66",
        "IEC104-POLLING": "67",
        "DL476-POLLING": "68",
        "继电保护远方操作规约": "70",
        "大唐先一通讯规约": "71",
        "IEC-104(不含一致性校验)": "72",
        "广东私有计划规约": "73",
        "转发CDT": "101",
        "转发BJ9702": "102",
        "转发DISA": "103",
        "转发IEC101": "121",
        "转发IEC104": "124",
        "转发IEC104(接收计划)": "125",

    }
    # 如果 s 在字典中，返回对应值；否则返回 s 自身
    return mapping.get(s, s)

# 是/否转换成1/0
def yn_station_type(s):
    """中文/数字 统一映射"""
    if pd.isna(s):
        return ""
    s = full_clean(s)

    if s in ("是",):
        return "1"
    elif s in ("否",):
        return "0"
    elif s in ("10kV",):
        return "112871465677750278"
    return s

# ========== 全局配置区 ==========
# 输出结果目录：所有对比结果、日志都会保存到这里，自动创建
OUTPUT_DIR = "C:\\Users\\13303\\Desktop\\excel_compare\\compare_result"

TASKS = [

    # 任务1：配网馈线表实时库和达梦对比
    {
        "name": "配网馈线表实时库和达梦对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kxb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kxb_dm.xls",
        "col_map": {
            # "所属厂站": "st_name",
            "馈线ID号": "ID",
            "馈线名称": "NAME",
            # "图形名": "GRAPH_NAME",
        },
        "transform_funcs": {
            "馈线ID号": lambda x: trim_str_length(x, 19),
            "馈线名称": full_clean,
            # "所属厂站": full_clean,
        },
        "key_col": "馈线ID号",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务2：开关站实时库商用库对比
    {
        "name": "开关站实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kg_zb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kg_zb_dm.xls",
        "col_map": {
            "开关站ID号": "ID",
            "开关站名称": "NAME",
            "所属馈线": "feeder_name",
            "开关站类型": "COMBINED_TYPE",
        },
        "transform_funcs": {
            "开关站ID号": lambda x: trim_str_length(x, 19),
            "开关站名称": full_clean,
            "所属馈线": full_clean,
            "开关站类型": convert_station_type,
        },
        "key_col": "开关站ID号",
        "ignore_only_row": False   # 忽略独有行
    },

    # 任务3：开关实时库商用库对比
    {
        "name": "开关实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kgb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\kgb_dm.xls",
        "col_map": {
            "开关ID号": "ID",
            "开关名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "所属组合设备": "composite_switch_name",
            # "开关类型": "BRK_TYPE",
            "开关联络类型": "BRK_CONNECT_TYPE",
        },
        "transform_funcs": {
            "开关ID号": lambda x: trim_str_length(x, 19),
            "开关名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            "所属组合设备": full_clean,
            # "开关类型": convert_station_type,
            "开关联络类型": convert_station_type,
        },
        "key_col": "开关ID号",
        "ignore_only_row": False   # 忽略独有行
    },

    # 任务4：组合开关实时库商用库对比
    {
        "name": "组合开关实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\zhkgb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\zhkgb_dm.xls",
        "col_map": {
            "组合开关名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
        },
        "transform_funcs": {
            "组合开关名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
        },
        "key_col": "组合开关名称",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务5：刀闸实时库商用库对比
    {
        "name": "刀闸实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\dzb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\dzb_dm.xls",
        "col_map": {
            "刀闸ID号": "ID",
            "刀闸名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "所属组合设备": "composite_switch_name",
            "刀闸类型": "DISCR_TYPE",
        },
        "transform_funcs": {
            "刀闸ID号": lambda x: trim_str_length(x, 19),
            "刀闸名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            "所属组合设备": full_clean,
            "刀闸类型": convert_station_type,
        },
        "key_col": "刀闸ID号",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务6：接地刀闸实时库商用库对比
    {
        "name": "接地刀闸实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\jddzb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\jddzb_dm.xls",
        "col_map": {
            "接地刀闸ID号": "ID",
            "接地刀闸名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "所属组合设备": "composite_switch_name",
            # "刀闸类型": "DISCR_TYPE",
        },
        "transform_funcs": {
            "接地刀闸ID号": lambda x: trim_str_length(x, 19),
            "接地刀闸名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            "所属组合设备": full_clean,
            # "刀闸类型": convert_station_type,
        },
        "key_col": "接地刀闸ID号",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务7：配网母线实时库商用库对比
    {
        "name": "母线表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\mxb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\mxb_dm.xls",
        "col_map": {
            "母线ID号": "ID",
            "母线名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            # "所属组合设备": "composite_switch_name",
            # "刀闸类型": "DISCR_TYPE",
        },
        "transform_funcs": {
            "母线ID号": lambda x: trim_str_length(x, 19),
            "母线名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            # "所属组合设备": full_clean,
            # "刀闸类型": convert_station_type,
        },
        "key_col": "母线ID号",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务8：配网终端信息表实时库商用库对比
    {
        "name": "终端信息表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\zdxxb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\zdxxb_dm.xls",
        "col_map": {
            "终端ID": "ID",
            "终端名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "终端编号": "CODE",
            "是否通讯状态统计": "IF_STAT_STATIC",
            "终端类别": "TERM_TYPE",
            # "运行定值区号": "CUR_FIXED_AREA",
            # "所属组合设备": "composite_switch_name",
            # "刀闸类型": "DISCR_TYPE",
        },
        "transform_funcs": {
            "终端ID": lambda x: trim_str_length(x, 19),
            "终端名称": full_clean,
            "所属开关站": full_clean,
            "终端编号": full_clean,
            "是否通讯状态统计": yn_station_type,
            "终端类别": convert_station_type,
            # "运行定值区号": full_clean,
            "所属馈线": full_clean,
            # "所属组合设备": full_clean,
            # "刀闸类型": convert_station_type,
        },
        "key_col": "终端ID",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务9：保护节点表实时库商用库对比
    {
        "name": "保护节点表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\bhjdb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\bhjdb_dm.xls",
        "col_map": {
            "标识": "ID",
            "中文名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "电压类型ID": "BV_ID",
            "类型": "PNT_TYPE",
            "开关数目": "BRK_NUM",
            "相应开关1": "RELY_BRK1",
            "相应开关2": "RELY_BRK2",
        },
        "transform_funcs": {
            "标识": lambda x: trim_str_length(x, 19),
            "中文名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            "电压类型ID": yn_station_type,
            "类型": convert_station_type,
            "开关数目": full_clean,
            "相应开关1": full_clean,
            "相应开关2": full_clean,
        },
        "key_col": "标识",
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务10：测点遥测表实时库商用库对比
    {
        "name": "测点遥测表表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\cdycb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\cdycb_dm.xls",
        "col_map": {
            "中文名称": "NAME",
            "所属开关站": "combined_name",
            "所属馈线": "feeder_name",
            "预留整型值1": "DEFAULT_I1",
            "预留长整型2": "DEFAULT_L1",
        },
        "transform_funcs": {
            "中文名称": full_clean,
            "所属开关站": full_clean,
            "所属馈线": full_clean,
            "预留整型值1": convert_station_type,
            "预留长整型2": lambda x: trim_str_length(x, 19),
        },
        "key_col": ["中文名称","所属开关站"],
        "ignore_only_row": False   # 忽略独有行
    },
    # 任务11：断路器DA控制表实时库商用库对比
    {
        "name": "断路器DA控制表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\dlqdakzb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\dlqdakzb_dm.xls",
        "col_map": {
            "标识": "ID",
            "厂站名称": "st_name",
            "开关名称": "cb_name",
            "关连馈线": "feeder_name",
            "运行状态": "SIMU_MODE",
            "执行模式": "CTRL_MODE",
            "图形名称": "GRAPH_NAME",
            "故障启动条件": "FAULT_START",
        },
        "transform_funcs": {
            "标识": lambda x: trim_str_length(x, 19),
            "厂站名称": full_clean,
            "开关名称": full_clean,
            "关连馈线": full_clean,
            "运行状态": convert_station_type,
            "执行模式": convert_station_type,
            "故障启动条件": convert_station_type,
            "图形名称": full_clean,
        },
        "key_col": "标识",
        "ignore_only_row": False,   # 忽略独有行

        # ========== 新增：跨列处理配置 ==========
        # 新增：指定跨列模式为跨文件参考
        #  cross_mode == "same_file"同文件内参考，两个文件各自独立处理
        #  cross_mode == "cross_file" 文件1的目标列，参考文件2的参考列长度
        "cross_col_mode": "cross_file",
        "cross_col_transforms": [
            # target_col 左侧去除 ref_col 的长度
            # left_trim_by_ref_len(target_col="开关名称", ref_col="厂站名称"),
            # 可继续添加多个规则，格式同上
            # 示例：target_col 从右侧保留 ref_col 长度的内容
            right_keep_by_ref_len(target_col="开关名称", ref_col="开关名称"),
        ]
    },
    # 任务12：配网通道表实时库商用库对比
    {
        "name": "配网通道表实时库商用库对比",
        "file1": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\tdb.xls",
        "file2": "C:\\Users\\13303\\Desktop\\excel_compare\\data\\tdb_dm.xls",
        "col_map": {
            "通道编号": "CHAN_NO",
            "通道名称": "CHAN_NAME",
            "通讯终端ID": "com_terminal_name",
            "网络描述一": "NET_DESCRIPTION1",
            "通信规约类型": "PROTO_TYPE",
            # "所属系统": "PARA_8",
        },
        "transform_funcs": {
            "通道编号": full_clean,
            "通道名称": full_clean,
            "通讯终端ID": full_clean,
            "网络描述一": full_clean,
            "通信规约类型": convert_station_type,
            # "所属系统": convert_station_type,
        },
        "key_col": "通道编号",
        "ignore_only_row": False   # 忽略独有行
    },
]

# ========== 批量任务配置区 ==========
# ignore_only_row: True=忽略独有行，只对比共同数据；False=严格校验行数一致


# ========== 文件读取 ==========
def load_excel(file_path):
    if not os.path.exists(file_path):
        print(f"错误：文件不存在 {file_path}")
        return None

    suffix = os.path.splitext(file_path)[1].lower()
    try:
        if suffix == ".xls":
            df = pd.read_excel(file_path, engine="xlrd")
        elif suffix == ".xlsx":
            df = pd.read_excel(file_path, engine="openpyxl")
        else:
            df = pd.read_excel(file_path)

        df.columns = [str(c).strip() for c in df.columns]
        print(f"【调试】{os.path.basename(file_path)} 实际表头：{list(df.columns)}")
        return df
    except Exception as e1:
        # 兼容伪Excel(TSV)，多编码+制表符分隔
        try:
            enc_list = ["utf-8", "gb18030", "gbk"]
            for enc in enc_list:
                try:
                    df = pd.read_csv(file_path, encoding=enc, sep="\t")
                    df.columns = [str(c).strip() for c in df.columns]
                    print(f"【调试】使用编码 {enc} 读取成功")
                    print(f"【调试】{os.path.basename(file_path)} 实际表头：{list(df.columns)}")
                    return df
                except:
                    continue
            return None
        except Exception as e2:
            print(f"读取失败，详细错误：")
            traceback.print_exc()
            return None


# ========== 单任务对比逻辑 ==========
def compare_single_task(task):
    print("\n" + "=" * 70)
    print(f"【任务：{task['name']}】")
    print(f"对比文件：{task['file1']} vs {task['file2']}")
    print("=" * 70)

    try:
        df1 = load_excel(task["file1"])
        df2 = load_excel(task["file2"])
        if df1 is None or df2 is None:
            return False

        # 校验列是否存在
        map_keys = list(task["col_map"].keys())
        map_vals = list(task["col_map"].values())
        missing1 = [c for c in map_keys if c not in df1.columns]
        missing2 = [c for c in map_vals if c not in df2.columns]
        if missing1 or missing2:
            if missing1:
                print(f"❌ 文件1缺失列：{missing1}")
            if missing2:
                print(f"❌ 文件2缺失列：{missing2}")
            return False

        # 筛选列 + 统一列名（全部统一为文件1的列名，这样后面的transform才能找到列）
        df1 = df1[map_keys]  # 文件1的列名不变！
        rev_map = {v: k for k, v in task["col_map"].items()}
        df2 = df2[map_vals].rename(columns=rev_map)  # 只改文件2的列名，改成文件1的

        # 应用清洗/转换函数
        for col, func in task["transform_funcs"].items():
            df1[col] = df1[col].apply(func)
            df2[col] = df2[col].apply(func)

        # ========== 新增：跨列转换处理（多任务复用）==========
        # cross_transforms = task.get("cross_col_transforms", [])
        # for trans_func in cross_transforms:
        #     # 对两个文件执行完全相同的跨列处理
        #     df1 = df1.apply(trans_func, axis=1)
        #     # df2 = df2.apply(trans_func, axis=1)

        # cross_transforms = task.get("cross_col_transforms", [])
        # for trans_func in cross_transforms:
        #     # 从处理函数中提取目标列，通用适配
        #     target_col = getattr(trans_func, "target_col", None)
        #     if target_col and target_col in df1.columns and target_col in df2.columns:
        #         df1[target_col] = df1.apply(trans_func, axis=1)
        #         # df2[target_col] = df2.apply(trans_func, axis=1)
        #     else:
        #         print(f"⚠️ 跨列处理警告：未找到目标列{target_col}或处理函数无标记，跳过该规则")


        # ========== 跨列转换处理（支持同文件/跨文件两种模式）==========
        print("【调试】df1列名：", list(df1.columns))
        print("【调试】df2列名：", list(df2.columns))
        cross_transforms = task.get("cross_col_transforms", [])
        # 模式：same_file=同文件内参考（默认），cross_file=文件1参考文件2的列
        cross_mode = task.get("cross_col_mode", "same_file")

        for trans_func in cross_transforms:
            target_col = getattr(trans_func, "target_col", None)
            ref_col = getattr(trans_func, "ref_col", None)

            if not target_col or target_col not in df1.columns:
                print(f"⚠️ 跨列处理警告：目标列 {target_col} 不存在，跳过该规则")
                continue

            if cross_mode == "same_file":
                # 原有逻辑：同文件内参考，两个文件各自独立处理
                if ref_col not in df1.columns or ref_col not in df2.columns:
                    print(f"⚠️ 跨列处理警告：参考列 {ref_col} 不存在，跳过该规则")
                    continue
                df1[target_col] = df1.apply(trans_func, axis=1)
                # df2[target_col] = df2.apply(trans_func, axis=1)

            elif cross_mode == "cross_file":
                # 新增逻辑：文件1的目标列，参考文件2的参考列长度
                if ref_col not in df2.columns:
                    print(f"⚠️ 跨列处理警告：文件2中参考列 {ref_col} 不存在，跳过该规则")
                    continue

                # 按索引对齐，计算文件2参考列的字符长度
                ref_len_series = df2[ref_col].astype(str).str.strip().str.len()

                # 执行右侧保留截取
                def _cross_right_process(row):
                    target_str = str(row[target_col]).strip() if pd.notna(row[target_col]) else ""
                    keep_len = int(row["_temp_ref_len"])
                    if keep_len <= 0:
                        return ""
                    if len(target_str) <= keep_len:
                        return target_str
                    return target_str[-keep_len:]

                df1["_temp_ref_len"] = ref_len_series
                df1[target_col] = df1.apply(_cross_right_process, axis=1)
                df1.drop(columns=["_temp_ref_len"], inplace=True)


##
        key_col = task["key_col"]
        df1 = df1.set_index(key_col).sort_index()
        df2 = df2.set_index(key_col).sort_index()

        # 独有行
        only1 = sorted(set(df1.index) - set(df2.index))
        only2 = sorted(set(df2.index) - set(df1.index))

        if len(only1) > 0:
            print(f"\n❌ 仅在第一个文件的主键({len(only1)}条)：")
            for k in only1[:10]:
                print(f"  {k}")
            if len(only1) > 10:
                print(f"  ... 省略剩余 {len(only1)-10} 条")

        if len(only2) > 0:
            print(f"\n❌ 仅在第二个文件的主键({len(only2)}条)：")
            for k in only2[:10]:
                print(f"  {k}")
            if len(only2) > 10:
                print(f"  ... 省略剩余 {len(only2)-10} 条")

        # 共同行对比
        common_idx = df1.index.intersection(df2.index)
        df_common1 = df1.loc[common_idx]
        df_common2 = df2.loc[common_idx]

        # diff_df = df_common1.compare(df_common2, keep_shape=True, keep_equal=True)
        diff_df = df_common1.compare(df_common2, keep_shape=True)
        diff_count = diff_df.notna().any(axis=1).sum()

        if diff_count == 0:
            print(f"\n✅ 共同行【所有字段完全一致】")
        else:
            print(f"\n⚠️ 共同行存在 {diff_count} 处真实数据差异：")
            print("-" * 70)
            print(diff_df.dropna(how='all'))

        # 统计信息
        print(f"\n【统计】")
        print(f"  文件1总行数：{len(df1)}")
        print(f"  文件2总行数：{len(df2)}")
        print(f"  共同行数：{len(common_idx)}")
        print(f"  独有行数：文件1:{len(only1)} | 文件2:{len(only2)}")
        print(f"  数据差异行数：{diff_count}")





        # ========== 新增：导出当前任务结果到Excel ==========

        # 清洗任务名称，规避Windows非法文件名
        task_safe_name = task['name'].replace("\\", "").replace("/", "").replace(":", "").replace("*", "").replace("?",
                                                                                                                   "").replace(
            "\"", "").replace("<", "").replace(">", "").replace("|", "")
        excel_name = f"{task_safe_name}_对比结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = os.path.join(OUTPUT_DIR, excel_name)

        # 组装数据表
        df_only1 = pd.DataFrame({"仅文件1主键": only1})
        df_only2 = pd.DataFrame({"仅文件2主键": only2})
        df_diff = df_common1.join(df_common2, rsuffix="_文件2")
        df_stat = pd.DataFrame([{
            "文件1总行数": len(df1),
            "文件2总行数": len(df2),
            "共同行数": len(common_idx),
            "文件1独有行数": len(only1),
            "文件2独有行数": len(only2),
            "差异行数": diff_count
        }])

        # 多工作表写入Excel
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_diff.to_excel(writer, sheet_name="差异数据", index=True)
            df_only1.to_excel(writer, sheet_name="仅文件1独有行", index=False)
            df_only2.to_excel(writer, sheet_name="仅文件2独有行", index=False)
            df_stat.to_excel(writer, sheet_name="统计信息", index=False)
        print(f"✅ 任务结果已导出至：{excel_path}")



        # # 清洗任务名称，规避Windows非法文件名
        # task_safe_name = task['name'].replace("\\", "").replace("/", "").replace(":", "").replace("*", "").replace("?","").replace("\"", "").replace("<", "").replace(">", "").replace("|", "")
        # excel_name = f"{task_safe_name}_对比结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # excel_path = os.path.join(OUTPUT_DIR, excel_name)
        #
        # # 组装数据表
        # df_only1 = pd.DataFrame({"仅文件1主键": only1})
        # df_only2 = pd.DataFrame({"仅文件2主键": only2})
        # df_diff = df_common1.join(df_common2, rsuffix="_文件2")
        # df_stat = pd.DataFrame([{
        #     "文件1总行数": len(df1),
        #     "文件2总行数": len(df2),
        #     "共同行数": len(common_idx),
        #     "文件1独有行数": len(only1),
        #     "文件2独有行数": len(only2),
        #     "差异行数": diff_count
        # }])
        #
        # # 第一步：先写入所有Sheet数据
        # with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        #     df_diff.to_excel(writer, sheet_name="差异数据", index=True)
        #     df_only1.to_excel(writer, sheet_name="仅文件1独有行", index=False)
        #     df_only2.to_excel(writer, sheet_name="仅文件2独有行", index=False)
        #     df_stat.to_excel(writer, sheet_name="统计信息", index=False)
        #
        # # 2. 统一加载工作簿（保证wb变量一定被定义）
        # from openpyxl import load_workbook
        # diff_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        # wb = load_workbook(excel_path)
        # ws = wb["差异数据"]
        #
        # # 3. 仅存在差异时，执行单元格标色
        # if diff_count > 0:
        #     # 获取pandas原生差异结果，保持判定规则统一
        #     raw_diff = df_common1.compare(df_common2, keep_shape=True)
        #     diff_index_list = raw_diff.index.tolist()
        #
        #     max_r = ws.max_row
        #     max_c = ws.max_column
        #     # 跳过表头，遍历数据行
        #     for row_idx in range(2, max_r + 1):
        #         row_key = ws.cell(row=row_idx, column=1).value
        #         if row_key not in diff_index_list:
        #             continue
        #         # 成对列比对标色
        #         for col in range(2, max_c, 2):
        #             cell_left = ws.cell(row=row_idx, column=col)
        #             cell_right = ws.cell(row=row_idx, column=col + 1)
        #             val1 = str(cell_left.value).strip() if cell_left.value is not None else ""
        #             val2 = str(cell_right.value).strip() if cell_right.value is not None else ""
        #             if val1 != val2:
        #                 cell_left.fill = diff_fill
        #                 cell_right.fill = diff_fill
        #
        # # 4. 统一保存文件（无论有无差异都会执行）
        # wb.save(excel_path)
        #
        # # 5. 输出提示
        # if diff_count == 0:
        #     print(f"✅ 任务结果已导出至：{excel_path}（无数据差异，未标记单元格）")
        # else:
        #     print(f"✅ 任务结果已导出至：{excel_path}（仅差异单元格已高亮）")


        # 判定任务成功
        ignore_only = task.get("ignore_only_row", False)
        if ignore_only:
            # 只看共同行是否无差异
            return diff_count == 0
        else:
            # 严格模式：无独有行 + 无差异
            return (len(only1) == 0 and len(only2) == 0 and diff_count == 0)

    except Exception as e:
        print(f"❌ 任务执行异常：{str(e)}")
        traceback.print_exc()
        return False


def main():
    #  自动创建结果目录
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    # 生成日志文件名（时间戳）
    log_name = f"compare_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    log_path = os.path.join(OUTPUT_DIR, log_name)
    # 打开日志文件，绑定双输出
    log_file = open(log_path, "w", encoding="utf-8")
    sys.stdout = TeeOutput(sys.stdout, log_file)

    # 对比文件
    print("=" * 70)
    print("批量Excel对比工具（深度清洗+区分独有行）")
    print(f"共加载 {len(TASKS)} 个对比任务")
    print("=" * 70)

    success_count = 0
    for idx, task in enumerate(TASKS):
        print(f"\n[{idx+1}/{len(TASKS)}] 开始执行：{task['name']}")
        ok = compare_single_task(task)
        if ok:
            success_count += 1

    print("\n" + "=" * 70)
    print(f"执行汇总：成功 {success_count} / 总计 {len(TASKS)}")
    print("=" * 70)

    # 关闭日志文件
    log_file.close()
    # 恢复标准输出
    sys.stdout = sys.__stdout__


if __name__ == "__main__":
    main()